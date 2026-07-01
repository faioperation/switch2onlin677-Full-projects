"""
utils/excel_styles.py
=====================
Shared Excel styling constants for the Dhifaf Baghdad export system.

Column definitions have MOVED to schemas/export_columns.py so both the
export generator and the upload validator can import them from a single
source of truth.  This module re-exports EXPORT_COLUMNS and TOTAL_COLS
for backward compatibility.
"""
from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from schemas.export_columns import PRODUCT_IMPORT_EXPORT_COLUMNS  # noqa: F401

# ── Backward-compat aliases used by excel_export_service.py ──────────────────
EXPORT_COLUMNS = PRODUCT_IMPORT_EXPORT_COLUMNS
TOTAL_COLS: int = len(EXPORT_COLUMNS)

# ── Color palette (hex strings, no leading #) ─────────────────────────────────
C_HEADER_BG       = "1E3A5F"   # Dark navy          — import-compatible headers
C_HEADER_READONLY = "3D5A7A"   # Medium navy        — export-only / read-only headers
C_HEADER_FG       = "FFFFFF"   # White
C_TITLE_BG        = "0D2137"   # Near-black navy
C_TITLE_FG        = "F5C518"   # Gold
C_META_BG         = "EBF0F7"   # Light blue-grey
C_SEP_BG          = "CDD8E8"   # Steel blue separator
C_ROW_ODD         = "F4F7FB"   # Very light stripe
C_ROW_EVEN        = "FFFFFF"   # White
C_IN_STOCK        = "E8F5E9"   # Soft green
C_LOW_STOCK       = "FFF8E1"   # Amber (qty 1–5)
C_OUT_STOCK       = "FFEBEE"   # Soft red
C_BORDER          = "BCC8D8"   # Steel blue-grey

# ── Border sides ──────────────────────────────────────────────────────────────
_THIN   = Side(border_style="thin",   color=C_BORDER)
_MEDIUM = Side(border_style="medium", color="8FA8C8")

BORDER_CELL   = Border(left=_THIN,   right=_THIN,   top=_THIN,   bottom=_THIN)
BORDER_HEADER = Border(left=_MEDIUM, right=_MEDIUM, top=_MEDIUM, bottom=_MEDIUM)
BORDER_NONE   = Border()

# ── Alignments ────────────────────────────────────────────────────────────────
ALIGN_TITLE   = Alignment(horizontal="left",   vertical="center", wrap_text=False, indent=1)
ALIGN_META    = Alignment(horizontal="left",   vertical="center", wrap_text=False, indent=1)
ALIGN_HEADER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=False)
ALIGN_LEFT    = Alignment(horizontal="left",   vertical="center", wrap_text=False)
ALIGN_RIGHT   = Alignment(horizontal="right",  vertical="center", wrap_text=False)

# ── Fills ─────────────────────────────────────────────────────────────────────
def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

FILL_TITLE        = _fill(C_TITLE_BG)
FILL_HEADER       = _fill(C_HEADER_BG)        # import-compatible column headers
FILL_HEADER_RO    = _fill(C_HEADER_READONLY)  # export-only column headers
FILL_META         = _fill(C_META_BG)
FILL_SEP          = _fill(C_SEP_BG)
FILL_ROW_ODD      = _fill(C_ROW_ODD)
FILL_ROW_EVEN     = _fill(C_ROW_EVEN)
FILL_IN_STOCK     = _fill(C_IN_STOCK)
FILL_LOW_STOCK    = _fill(C_LOW_STOCK)
FILL_OUT_STOCK    = _fill(C_OUT_STOCK)

# ── Fonts ─────────────────────────────────────────────────────────────────────
FONT_TITLE    = Font(name="Calibri", bold=True,  size=14, color=C_TITLE_FG)
FONT_META_KEY = Font(name="Calibri", bold=True,  size=10, color="2D3748")
FONT_META_VAL = Font(name="Calibri", bold=False, size=10, color="4A5568")
FONT_HEADER   = Font(name="Calibri", bold=True,  size=10, color=C_HEADER_FG)
FONT_DATA     = Font(name="Calibri", bold=False, size=10, color="1A202C")

# ── Number formats ────────────────────────────────────────────────────────────
# Re-exported from export_columns for callers that import from here.
from schemas.export_columns import (   # noqa: E402, F401
    FMT_DECIMAL,
    FMT_INT,
    FMT_IQD,
    FMT_PCT,
    FMT_TEXT,
    FMT_USD,
)
